"""Tool: Generate Excel, Markdown, and Log outputs."""

import os
import datetime
from typing import List, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill


def generate_excel(results: List[Dict], summary: Dict, output_path: str):
    """Generate formatted Excel report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison Report"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Headers
    headers = ["Line #", "Base Text", "Compare Text", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for i, r in enumerate(results, 2):
        ws.cell(row=i, column=1, value=r["line_number"])
        ws.cell(row=i, column=2, value=r["base_text"])
        ws.cell(row=i, column=3, value=r["compare_text"])
        status_cell = ws.cell(row=i, column=4, value=r["status"])

        if r["status"] == "PASS":
            status_cell.fill = green_fill
        else:
            status_cell.fill = red_fill

    # Auto-column width
    for col in range(1, 5):
        max_len = len(str(ws.cell(row=1, column=col).value))
        for row in range(2, len(results) + 2):
            val = str(ws.cell(row=row, column=col).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 2, 80)

    wb.save(output_path)
    wb.close()


def generate_markdown(results: List[Dict], summary: Dict, output_path: str):
    """Generate Markdown summary report."""
    lines = []
    lines.append("# File Comparison Report")
    lines.append(f"*Generated: {datetime.datetime.now().isoformat()}*")
    lines.append("")
    lines.append("## Summary")
    lines.append(f"- **Total Lines:** {summary['total_lines']}")
    lines.append(f"- **Pass:** {summary['pass_count']}")
    lines.append(f"- **Fail:** {summary['fail_count']}")
    lines.append(f"- **Pass Rate:** {summary['pass_percentage']}%")
    lines.append("")
    lines.append("## Detailed Results")
    lines.append("")
    lines.append("| Line # | Base Text | Compare Text | Status |")
    lines.append("|--------|-----------|--------------|--------|")

    for r in results:
        base = r["base_text"].replace("|", "\\|")[:60]
        comp = r["compare_text"].replace("|", "\\|")[:60]
        lines.append(f"| {r['line_number']} | {base} | {comp} | {r['status']} |")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def generate_log(results: List[Dict], summary: Dict, base_name: str, comp_name: str, output_path: str):
    """Generate detailed execution log."""
    now = datetime.datetime.now()
    lines = []
    lines.append("=" * 60)
    lines.append(f"FILE COMPARISON LOG")
    lines.append(f"Timestamp: {now.isoformat()}")
    lines.append("=" * 60)
    lines.append("")
    lines.append(f"[{now.strftime('%H:%M:%S')}] Comparison started")
    lines.append(f"[{now.strftime('%H:%M:%S')}] Base file: {base_name}")
    lines.append(f"[{now.strftime('%H:%M:%S')}] Compare file: {comp_name}")
    lines.append(f"[{now.strftime('%H:%M:%S')}] Total lines compared: {summary['total_lines']}")
    lines.append(f"[{now.strftime('%H:%M:%S')}] Pass count: {summary['pass_count']}")
    lines.append(f"[{now.strftime('%H:%M:%S')}] Fail count: {summary['fail_count']}")
    lines.append(f"[{now.strftime('%H:%M:%S')}] Pass rate: {summary['pass_percentage']}%")
    lines.append("")

    # Detailed trace
    fail_count = 0
    for r in results:
        if r["status"] == "FAIL":
            fail_count += 1
            lines.append(f"[FAIL] Line {r['line_number']}:")
            lines.append(f"       Base:    {r['base_text'][:100]}")
            lines.append(f"       Compare: {r['compare_text'][:100]}")
            lines.append("")

    if fail_count == 0:
        lines.append("[INFO] No failures detected — all lines PASS.")
    else:
        lines.append(f"[INFO] Total failures: {fail_count}")

    lines.append("")
    lines.append("=" * 60)
    lines.append("END OF LOG")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def generate_all(results: List[Dict], summary: Dict, base_name: str, comp_name: str, output_dir: str):
    """Generate all three report files."""
    excel_path = os.path.join(output_dir, "report.xlsx")
    md_path = os.path.join(output_dir, "report.md")
    log_path = os.path.join(output_dir, "report.log")

    generate_excel(results, summary, excel_path)
    generate_markdown(results, summary, md_path)
    generate_log(results, summary, base_name, comp_name, log_path)

    return {
        "excel": excel_path,
        "markdown": md_path,
        "log": log_path,
    }
