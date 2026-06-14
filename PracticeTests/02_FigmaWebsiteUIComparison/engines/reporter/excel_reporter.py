"""Excel Reporter — exports comparison results to .xlsx."""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from loggers.logger import get_logger

logger = get_logger(__name__)


class ExcelReporter:
    """Generates Excel reports from comparison results."""

    @staticmethod
    def generate(session: dict, session_dir: str, output_dir: str, elements: list = None) -> str:
        """Generate Excel report with 3 sheets.

        Args:
            session: Session dict.
            session_dir: Path to session data.
            output_dir: Output directory.
            elements: Pre-loaded element list with checks. If None, loads from session.

        Returns:
            Path to generated .xlsx file.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()

        summary = session.get("summary", {}) or {}
        config = session.get("config", {}) or {}

        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.append(["Visual UI Testing Platform — Report"])
        ws1.append([])
        ws1.append(["Session ID", session.get("id", "")])
        ws1.append(["Figma URL", config.get("figma_url", "")])
        ws1.append(["Web URL", config.get("web_url", "")])
        ws1.append(["Created", session.get("created_at", "")])
        ws1.append(["Status", session.get("status", "")])
        ws1.append([])
        ws1.append(["Overall Similarity", f"{summary.get('overall_similarity', 0)}%"])
        ws1.append(["Total Elements", summary.get("total_elements", 0)])
        ws1.append(["Total Checks", summary.get("total_checks", 0)])
        ws1.append(["Pass Count", summary.get("pass_count", 0)])
        ws1.append(["Fail Count", summary.get("fail_count", 0)])
        ws1.append(["Verdict", summary.get("verdict", "")])
        ws1.append([])
        ws1.append(["Severity Breakdown"])
        by_severity = summary.get("by_severity", {}) or {}
        for sev in ["critical", "high", "medium", "low"]:
            ws1.append([sev.capitalize(), by_severity.get(sev, 0)])
        ws1.append([])
        ws1.append(["Category Breakdown"])
        by_category = summary.get("by_category", {}) or {}
        for cat, data in (by_category or {}).items():
            pct = data.get("pass_pct", 0) if data else 0
            ws1.append([cat.capitalize(), f"{pct}% pass"])

        # Style header
        header_font = Font(bold=True, size=14)
        ws1["A1"].font = header_font

        # Sheet 2: All Checks
        ws2 = wb.create_sheet("All Checks")
        ws2.append(["Element", "Category", "Property", "Expected", "Actual", "Unit", "Status", "Severity", "Difference"])
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 18
        ws2.column_dimensions["C"].width = 22
        ws2.column_dimensions["D"].width = 25
        ws2.column_dimensions["E"].width = 25
        ws2.column_dimensions["F"].width = 10
        ws2.column_dimensions["G"].width = 10
        ws2.column_dimensions["H"].width = 12
        ws2.column_dimensions["I"].width = 15

        red_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        green_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")

        if elements:
            for elem in elements:
                elem_name = elem.get("figma_element", {}).get("name", "Unknown") if elem.get("figma_element") else "Unknown"
                checks = elem.get("checks", [])
                if not checks:
                    ws2.append([elem_name, "", "", "", "", "", "PASS", "none", ""])
                for check in checks:
                    row = [
                        elem_name,
                        check.get("category", ""),
                        check.get("property", ""),
                        str(check.get("expected", "")),
                        str(check.get("actual", "")),
                        check.get("unit", ""),
                        check.get("status", ""),
                        check.get("severity", ""),
                        str(check.get("difference", "")),
                    ]
                    ws2.append(row)
                    row_num = ws2.max_row
                    if check.get("status") == "FAIL":
                        for col in range(1, 10):
                            ws2.cell(row=row_num, column=col).fill = red_fill
                    else:
                        for col in range(1, 10):
                            ws2.cell(row=row_num, column=col).fill = green_fill

        # Sheet 3: Summary data
        ws3 = wb.create_sheet("By Category")
        ws3.append(["Category", "Pass", "Fail", "Pass %"])
        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 10
        ws3.column_dimensions["C"].width = 10
        ws3.column_dimensions["D"].width = 10

        for cat, data in (by_category or {}).items():
            ws3.append([
                cat.capitalize(),
                data.get("pass", 0) if data else 0,
                data.get("fail", 0) if data else 0,
                f"{data.get('pass_pct', 0):.1f}%" if data else "0%",
            ])

        path = os.path.join(output_dir, "report.xlsx")
        wb.save(path)
        logger.info("Excel report saved: %s", path)
        return path
